"""Lê registro_offline_view.xlsx (monitoramento de equipamentos) e a API DATATEM
(status/consumo de SIM cards) e gera um dashboard estático (index.html) com dois painéis.

Uso:
  1. cp .env.example .env  (preencha com suas credenciais DATATEM)
  2. Coloque seu registro_offline_view.xlsx e monitor_fundo.log na raiz do projeto
     (ou ajuste XLSX/LOG abaixo)
  3. python gerar.py  ->  abre index.html no navegador

Rode de novo pra atualizar depois que o monitor gravar mais dados.
"""
import base64
import json
import os
import re
import time
from collections import Counter, defaultdict
from datetime import datetime, timedelta

import requests
from dotenv import dotenv_values

AQUI = os.path.dirname(__file__)
XLSX = os.path.join(AQUI, "registro_offline_view.xlsx")
LOG = os.path.join(AQUI, "monitor_fundo.log")
SAIDA = os.path.join(AQUI, "index.html")
ENV_CHIPS = os.path.join(AQUI, ".env")
TOKEN_CACHE = os.path.join(AQUI, ".token_cache.json")
HIST_CHIPS = os.path.join(AQUI, "chips_historico.jsonl")
RESET_LOG = os.path.join(AQUI, "chips_reset_historico.jsonl")
FMT = "%d-%m-%Y %H:%M:%S"
DIVISOR_RE = re.compile(r"-{2,}\s*(\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}:\d{2})\s*-{2,}")
TS_RE = re.compile(r"^\[(\d{2}-\d{2}-\d{4} \d{2}:\d{2}:\d{2})\]", re.MULTILINE)
GAP_LIMITE_MIN = 10  # nota: acima disso considera que o script de monitoramento parou


def carregar():
    from openpyxl import load_workbook
    ws = load_workbook(XLSX, read_only=True).active
    linhas, ciclos = [], set()  # ciclos = todos os timestamps de ciclo (divisores)
    for codigo, data_desat, campo, registrada, reset in ws.iter_rows(min_row=2, values_only=True):
        if codigo is None:
            continue
        m = DIVISOR_RE.match(str(codigo))
        if m:
            ciclos.add(m.group(1))
            continue
        if registrada:
            linhas.append((codigo, registrada, str(reset or "")))
    return linhas, ciclos


def agregar(linhas, ciclos):
    por_codigo = defaultdict(lambda: {"ciclos": 0, "resets": 0, "primeiro": None, "ultimo": None})
    # nota: comparar os dt parseados, não a string "registrada" direto - "DD-MM-YYYY" como
    # string ordena errado entre meses/dias diferentes (bug real encontrado 2026-08-03: "01-08"
    # perdia pra "31-07" porque '0' < '3' no primeiro caractere, mesmo Ago vindo depois de Jul).
    primeiro_dt, ultimo_dt = {}, {}
    por_ciclo = defaultdict(set)  # registrada -> set de codigos offline naquele ciclo
    for codigo, registrada, reset in linhas:
        d = por_codigo[codigo]
        d["ciclos"] += 1
        if reset.strip().lower() == "sim":
            d["resets"] += 1
        dt = datetime.strptime(registrada, FMT)
        if codigo not in primeiro_dt or dt < primeiro_dt[codigo]:
            primeiro_dt[codigo] = dt
            d["primeiro"] = registrada
        if codigo not in ultimo_dt or dt > ultimo_dt[codigo]:
            ultimo_dt[codigo] = dt
            d["ultimo"] = registrada
        por_ciclo[registrada].add(codigo)

    ranking = sorted(
        ({"codigo": c, **v} for c, v in por_codigo.items()),
        key=lambda x: x["ciclos"], reverse=True,
    )

    # eixo do tempo = união de divisores + ciclos com offline, ordenado cronologicamente
    todos = sorted(ciclos | set(por_ciclo), key=lambda s: datetime.strptime(s, FMT))
    timeline = [
        {"t": ts, "n": len(por_ciclo.get(ts, ())), "codigos": sorted(por_ciclo.get(ts, ()))}
        for ts in todos
    ]
    pico = max((p["n"] for p in timeline), default=0)
    # quantos ciclos tiveram N simultâneos (distribuição da simultaneidade)
    dist = Counter(p["n"] for p in timeline if p["n"] > 0)

    return {
        "gerado_em": datetime.now().strftime(FMT),
        "ranking": ranking,
        "timeline": timeline,
        "dist": [{"n": n, "ciclos": dist[n]} for n in sorted(dist)],
        "totais": {
            "equipamentos": len(por_codigo),
            "ciclosOffline": sum(v["ciclos"] for v in por_codigo.values()),
            "resets": sum(v["resets"] for v in por_codigo.values()),
            "pico": pico,
        },
    }


def fmt_horas(h):
    """8d 16h 12m / 3d 16h / 16h 12m / 16h - sem decimais confusas (16.2h != "16h e 2min")."""
    total_min = round(h * 60)
    d, resto_min = divmod(total_min, 24 * 60)
    hh, mm = divmod(resto_min, 60)
    partes = ([f"{d}d"] if d else []) + [f"{hh}h"] + ([f"{mm}m"] if mm else [])
    return " ".join(partes)


def calcular_perdido(timestamps, limite_min=GAP_LIMITE_MIN):
    """Soma os intervalos entre ciclos maiores que limite_min (= o script parou de rodar)."""
    ts = sorted(set(timestamps))
    if len(ts) < 2:
        return timedelta(0), timedelta(0), None, None
    limite = timedelta(minutes=limite_min)
    perdido = sum((b - a for a, b in zip(ts, ts[1:]) if b - a > limite), timedelta(0))
    return perdido, ts[-1] - ts[0], ts[0], ts[-1]


def tempo_observacao():
    try:
        with open(LOG, encoding="utf-8", errors="replace") as f:
            texto = f.read()
    except FileNotFoundError:
        return {"desde": None, "observado": "0.0h", "perdido": "0.0h", "perdidoHoras": 0}
    timestamps = (datetime.strptime(t, FMT) for t in TS_RE.findall(texto))
    perdido, total, inicio, fim = calcular_perdido(timestamps)
    observado = total - perdido
    return {
        "desde": inicio.strftime(FMT) if inicio else None,
        "observado": fmt_horas(observado.total_seconds() / 3600),
        "perdido": fmt_horas(perdido.total_seconds() / 3600),
        "perdidoHoras": round(perdido.total_seconds() / 3600, 2),
    }


EXCLUIVEIS = set()  # nota: códigos a excluir da visão "filtrada" (ex: {"EQP-16"})


def _login_datatem(env, base_url):
    senha_b64 = base64.b64encode(env["DATATEM_PASSWORD"].encode()).decode()
    resp = requests.post(
        f"{base_url}/authorization/api/v1/auth/login",
        json={
            "username": env["DATATEM_USER"], "password": senha_b64,
            "grantType": "password", "clientId": env["DATATEM_CLIENT_ID"],
            "clientSecret": env["DATATEM_CLIENT_SECRET"],
        }, headers={"User-Agent": "Mozilla/5.0"}, timeout=15,
    ).json()
    return resp["authenticationToken"]


def _refresh_datatem(base_url, refresh_token):
    resp = requests.post(
        f"{base_url}/authorization/api/v1/auth/refresh_token",
        json={"refreshToken": refresh_token, "grantType": "refresh_token"},
        headers={"User-Agent": "Mozilla/5.0"}, timeout=15,
    )
    resp.raise_for_status()
    return resp.json()


def _obter_token(env, base_url):
    """Reaproveita o token salvo em cache; só loga de novo quando expira (ou refresh falha)."""
    cache = {}
    try:
        with open(TOKEN_CACHE, encoding="utf-8") as f:
            cache = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        pass

    agora = time.time()
    if cache.get("accessToken") and cache.get("expiresAt", 0) > agora + 60:
        return cache["accessToken"]

    if cache.get("refreshToken"):
        try:
            novo = _refresh_datatem(base_url, cache["refreshToken"])
            cache = {
                "accessToken": novo["accessToken"], "refreshToken": novo["refreshToken"],
                "expiresAt": agora + novo.get("expiresIn", 3600),
            }
            with open(TOKEN_CACHE, "w", encoding="utf-8") as f:
                json.dump(cache, f)
            return cache["accessToken"]
        except Exception:
            pass  # refresh token também expirou/inválido -> loga do zero abaixo

    auth = _login_datatem(env, base_url)
    cache = {
        "accessToken": auth["accessToken"], "refreshToken": auth["refreshToken"],
        "expiresAt": agora + auth.get("expiresIn", 3600),
    }
    with open(TOKEN_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f)
    return cache["accessToken"]


def _gravar_historico(lista, gerado_em):
    # nota: cresce ~1 linha/chip por execução; se virar gigabyte, rotacionar por mês.
    with open(HIST_CHIPS, "a", encoding="utf-8") as f:
        for x in lista:
            f.write(json.dumps({
                "ts": gerado_em, "iccid": x["iccid"], "codigo": x["codigo"], "cliente": x["cliente"],
                "status": x["status"], "conectado": x["conectado"], "pct": x["pct"],
            }, ensure_ascii=False) + "\n")


def carregar_chips():
    """Busca status de conexão + consumo de todos os simcards na API DATATEM."""
    env = dotenv_values(ENV_CHIPS)
    base_url = env.get("DATATEM_API_URL", "https://app-gateway.brcaptura.com.br")
    token = _obter_token(env, base_url)
    headers = {"Authorization": f"Bearer {token}", "User-Agent": "Mozilla/5.0"}

    chips, page = [], 1
    while True:
        resp = requests.get(
            f"{base_url}/simcard/connection/v1/connections",
            params={"pageSize": 100, "pageNumber": page}, headers=headers, timeout=15,
        )
        if resp.status_code in (401, 403) and page == 1:
            # nota: token em cache pode ter sido revogado no servidor mesmo sem ter expirado -> força login novo
            auth = _login_datatem(env, base_url)
            with open(TOKEN_CACHE, "w", encoding="utf-8") as f:
                json.dump({"accessToken": auth["accessToken"], "refreshToken": auth["refreshToken"],
                           "expiresAt": time.time() + auth.get("expiresIn", 3600)}, f)
            headers["Authorization"] = f"Bearer {auth['accessToken']}"
            continue
        r = resp.json()
        content = r.get("content", [])
        if not content:
            break
        chips.extend(content)
        if len(content) < 100:
            break
        page += 1

    def custom_field(c, nome):
        for v in (c.get("customField") or {}).get("values") or []:
            if v.get("name") == nome and v.get("value"):
                return v["value"]
        return None

    def horas_desde(iso_str):
        if not iso_str:
            return None
        dt = datetime.fromisoformat(iso_str)
        agora = datetime.now(dt.tzinfo) if dt.tzinfo else datetime.now()
        return (agora - dt).total_seconds() / 3600

    def status_bucket(horas):
        if horas is None:
            return "sem_comunicacao"
        if horas <= 2:
            return "atualizado"
        if horas <= 24:
            return "atrasado"
        return "sem_comunicacao"

    lista = []
    for c in chips:
        franquia = int(c.get("franchise") or 0)
        consumido = int(c.get("cycletotalbytes") or 0)
        horas = horas_desde(c.get("lastAccounting"))
        lista.append({
            "codigo": custom_field(c, "device") or c.get("simcard"),
            "cliente": (custom_field(c, "description2") or c.get("clientName") or "").strip().upper() or None,
            "operadora": c.get("operatorName"),
            "iccid": c.get("iccid"),
            "telefone": c.get("phonenumber"),
            "cidade": c.get("city"),
            "conectado": c.get("connectivitystatus") == "CONNECTED",
            "ultimaComunicacao": c.get("lastAccounting"),
            "status": status_bucket(horas),
            "critico": horas is None or horas > 48,
            "franquiaMb": round(franquia / 1_000_000, 1),
            "consumidoMb": round(consumido / 1_000_000, 1),
            "restanteMb": round(max(franquia - consumido, 0) / 1_000_000, 1),
            "pct": round(c.get("cycleTotalBytesPct") or 0, 1),
        })
    lista.sort(key=lambda x: x["pct"], reverse=True)
    conectados = sum(1 for x in lista if x["conectado"])
    _gravar_historico(lista, datetime.now().strftime(FMT))
    return {
        "lista": lista,
        "totais": {
            "chips": len(lista),
            "conectados": conectados,
            "desconectados": len(lista) - conectados,
            "atualizados": sum(1 for x in lista if x["status"] == "atualizado"),
            "atrasados": sum(1 for x in lista if x["status"] == "atrasado"),
            "semComunicacao": sum(1 for x in lista if x["status"] == "sem_comunicacao"),
            "criticos": sum(1 for x in lista if x["critico"]),
            "consumidoMb": round(sum(x["consumidoMb"] for x in lista), 1),
            "franquiaMb": round(sum(x["franquiaMb"] for x in lista), 1),
        },
    }


def carregar_ciclos_reset():
    """Cada linha do log é uma leva do reset automático (~1h). Retorna em ordem
    cronológica o que foi enviado em cada leva, pra download avulso na tela."""
    ciclos = []
    try:
        with open(RESET_LOG, encoding="utf-8") as f:
            linhas = f.readlines()
    except FileNotFoundError:
        return ciclos
    for linha in linhas:
        try:
            r = json.loads(linha)
        except json.JSONDecodeError:
            continue
        lote_falhou = "status_http" in r and r["status_http"] != 200
        ciclos.append({
            "ts": r.get("ts"),
            "enviados": [] if lote_falhou else r.get("enviados", []),
            "jaPendentes": r.get("ja_pendentes", []),
            "falhas": r.get("falhas", []),
        })
    return ciclos


def carregar_snapshots_chips():
    """Agrupa chips_historico.jsonl por ts de execução (cada carregar_chips() grava um grupo)."""
    grupos = defaultdict(list)
    try:
        with open(HIST_CHIPS, encoding="utf-8") as f:
            for linha in f:
                try:
                    r = json.loads(linha)
                except json.JSONDecodeError:
                    continue
                grupos[r["ts"]].append(r)
    except FileNotFoundError:
        pass
    return grupos


def anexar_snapshots(ciclos, snapshots):
    """Pra cada leva de reset, acha o snapshot de chips mais próximo (gravado pela
    mesma chamada de carregar_chips() que gerou aquela leva)."""
    tss_ordenados = sorted(snapshots, key=lambda s: datetime.strptime(s, FMT))
    for c in ciclos:
        alvo = datetime.strptime(c["ts"], FMT)
        candidatos = [t for t in tss_ordenados if datetime.strptime(t, FMT) <= alvo]
        c["snapshot"] = snapshots[candidatos[-1]] if candidatos else []


def calcular_eficacia_resets(ciclos, snapshots):
    """Pra cada reset enviado, descobre se o chip voltou a comunicar (status virou
    'atualizado' num snapshot) antes do próximo ciclo de reset - reaproveita os
    snapshots de alta frequência já gravados em chips_historico.jsonl, não precisa
    de nenhum dado novo. Marca cada item de "enviados" com reconectou (True/False/
    None - None = ciclo mais recente, ainda sem próximo ciclo pra comparar) e
    reconectouEm (ts ou None)."""
    tss_ordenados = sorted(snapshots, key=lambda s: datetime.strptime(s, FMT))
    por_ts = {t: {r["iccid"]: r for r in snapshots[t]} for t in tss_ordenados}
    for i, c in enumerate(ciclos):
        inicio = datetime.strptime(c["ts"], FMT)
        fim = datetime.strptime(ciclos[i + 1]["ts"], FMT) if i + 1 < len(ciclos) else None
        janela = [t for t in tss_ordenados
                  if datetime.strptime(t, FMT) > inicio and (fim is None or datetime.strptime(t, FMT) < fim)]
        for item in c.get("enviados", []):
            reconectou_em = next(
                (t for t in janela if por_ts[t].get(item["iccid"], {}).get("status") == "atualizado"), None)
            item["reconectou"] = None if fim is None else bool(reconectou_em)
            item["reconectouEm"] = reconectou_em


def carregar_reset_chips():
    """Lê o histórico de resets automáticos (reset_chips_datatem.py, script privado -
    não faz parte deste repo de exemplo). Retorna, por ICCID, o desfecho mais recente:
    reset confirmado (enviado) ou ainda aguardando. Registros antigos que marcavam o
    lote inteiro como enviado sem checar o HTTP de fato (campo status_http != 200 no
    nível do registro) são tratados como não enviados, mesmo que o campo "enviados"
    original diga o contrário."""
    estado = {}
    try:
        with open(RESET_LOG, encoding="utf-8") as f:
            linhas = f.readlines()
    except FileNotFoundError:
        return estado

    for linha in linhas:
        try:
            r = json.loads(linha)
        except json.JSONDecodeError:
            continue
        ts = r.get("ts")
        lote_falhou = "status_http" in r and r["status_http"] != 200

        def registrar(item, resultado):
            iccid = item["iccid"] if isinstance(item, dict) else item
            rotulo = item.get("estado") if isinstance(item, dict) else None
            estado[iccid] = {"quando": ts, "resultado": resultado, "estado": rotulo}

        for item in r.get("enviados", []):
            registrar(item, "aguardando" if lote_falhou else "enviado")
        for item in r.get("ja_pendentes", []):
            registrar(item, "aguardando")
        for item in r.get("falhas", []):
            registrar(item, "aguardando")
    return estado


def main():
    linhas, ciclos = carregar()
    completo = agregar(linhas, ciclos)
    filtrado = agregar([l for l in linhas if l[0] not in EXCLUIVEIS], ciclos)
    obs = tempo_observacao()
    try:
        chips = carregar_chips()
        reset_hist = carregar_reset_chips()
        for x in chips["lista"]:
            x["resetInfo"] = reset_hist.get(x["iccid"]) if x["status"] in ("atrasado", "sem_comunicacao") else None
        ciclos_reset = carregar_ciclos_reset()
        snapshots = carregar_snapshots_chips()
        anexar_snapshots(ciclos_reset, snapshots)
        calcular_eficacia_resets(ciclos_reset, snapshots)
        chips["ciclosReset"] = ciclos_reset
    except Exception as e:
        chips = {"erro": str(e)}
    payload = {"observacao": obs, "completo": completo, "filtrado": filtrado, "chips": chips}
    html = TEMPLATE.replace("__DADOS__", json.dumps(payload, ensure_ascii=False))
    with open(SAIDA, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Gerado: {SAIDA}")
    print(f"  {completo['totais']['equipamentos']} equipamentos, "
          f"{completo['totais']['ciclosOffline']} ciclos offline, "
          f"pico de {completo['totais']['pico']} simultâneos. "
          f"Observado {obs['observado']}, perdido {obs['perdido']}.")


# nota: Chart.js via CDN (a máquina do monitor está online). Se precisar 100%
# offline, baixar chart.umd.min.js pra pasta e trocar o src.
with open(os.path.join(AQUI, "template.html"), encoding="utf-8") as _f:
    TEMPLATE = _f.read()


if __name__ == "__main__":
    main()
